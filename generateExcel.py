from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Color, Font, Border, Side
from openpyxl.cell import Cell

wb = Workbook()

# grab the active worksheet
ws = wb.active

# Here we define the 3 row styles we can find through the calc ===================================================


# Data is the value of each cell
# isBold is a boolean indicating wether the last rows are in bold or bit
def styled_cells(value, isBold):
    index = 1
    for c in value:
        if index == 1:
            c = Cell(ws, column="A", row=1, value=c)
            c.font = Font(bold=True, size=9)
            c.border = Border(left=Side("thin"))
            c.alignment = Alignment(horizontal="center", vertical="center")
        if index == 2:
            c = Cell(ws, column="A", row=1, value=c)
            c.font = Font(bold=True, size=9)
            c.border = Border(left=Side("thin"))
            c.alignment = Alignment(horizontal="center", vertical="center")
        if index == 3:
            c = Cell(ws, column="A", row=1, value=c)
            c.font = Font(bold=isBold, size=9)
            c.alignment = Alignment(horizontal="center", vertical="center")
        if index == 4:
            c = Cell(ws, column="A", row=1, value=c)
            c.font = Font(bold=isBold, size=9)
            c.border = Border(left=Side("thin"))
            c.alignment = Alignment(horizontal="center", vertical="center")
        if index == 5:
            c = Cell(ws, column="A", row=1, value=c)
            c.font = Font(bold=isBold, size=9)
            c.border = Border(right=Side("thin"))
            c.alignment = Alignment(horizontal="center", vertical="center")
        index = index + 1
        yield c


def total_fatigue_stress_styled_cells(value):
    index = 1
    for c in value:
        if index == 1:
            c = Cell(ws, column="A", row=1, value=c)
            c.font = Font(bold=True, size=9)
            c.border = Border(left=Side("thin"), bottom=Side("thin"))
            c.alignment = Alignment(horizontal="center", vertical="center")
        if index == 2:
            c = Cell(ws, column="A", row=1, value=c)
            c.border = Border(bottom=Side("thin"))
            c.alignment = Alignment(horizontal="center", vertical="center")
        if index == 3:
            c = Cell(ws, column="A", row=1, value=c)
            c.border = Border(bottom=Side("thin"))
            c.alignment = Alignment(horizontal="center", vertical="center")
        if index == 4:
            c = Cell(ws, column="A", row=1, value=c)
            c.border = Border(bottom=Side("thin"))
            c.alignment = Alignment(horizontal="center", vertical="center")
        if index == 5:
            c = Cell(ws, column="A", row=1, value=c)
            c.border = Border(bottom=Side("thin"), right=Side("thin"))
            c.alignment = Alignment(horizontal="center", vertical="center")
        index = index + 1
        yield c


def bins_styled_cells(value):
    index = 1
    for c in value:
        if index == 1:
            c = Cell(ws, column="A", row=1, value=c)
            c.font = Font(bold=True, size=9)
            c.border = Border(left=Side("thin"), bottom=Side("thin"))
            c.alignment = Alignment(horizontal="center", vertical="center")
        if index == 2:
            c = Cell(ws, column="A", row=1, value=c)
            c.font = Font(bold=True, size=9)
            c.border = Border(left=Side("thin"), bottom=Side("thin"))
            c.alignment = Alignment(horizontal="center", vertical="center")
        if index == 3:
            c = Cell(ws, column="A", row=1, value=c)
            c.font = Font(bold=True, size=9)
            c.border = Border(bottom=Side("thin"))
            c.alignment = Alignment(horizontal="center", vertical="center")
        if index == 4:
            c = Cell(ws, column="A", row=1, value=c)
            c.font = Font(bold=True, size=9)
            c.border = Border(left=Side("thin"), bottom=Side("thin"))
            c.alignment = Alignment(horizontal="center", vertical="center")
        if index == 5:
            c = Cell(ws, column="A", row=1, value=c)
            c.font = Font(bold=True, size=9)
            c.border = Border(right=Side("thin"), bottom=Side("thin"))
            c.alignment = Alignment(horizontal="center", vertical="center")
        index = index + 1
        yield c


# Here is a fake array to test the rows insertions ================================================================

rows = [1, 2, 3] #TODO replace with real data

# Here we create the resulting document's headers ================================================================
ws.append(
    ["STRESS HISTORY BLABLABLA"]
)  # TODO: Change the headers name to correct values
# Here we center the text in the cell
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
# here we fuse all the first row's cell to create a header
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
# Here we add borders to the header's cells
for cell in ws["1"]:
    cell.border = Border(bottom=Side("thin"), left=Side("thin"), right=Side("thin"))

# Same process for the second header
ws.append(["WAVE & MOTION BLABLABLA"])
ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
# in this one we also filled the cells with a background color
ws["A2"].fill = PatternFill(
    patternType="solid", fill_type="solid", fgColor=Color("C4C4C4")
)
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
for cell in ws["2"]:
    cell.border = Border(bottom=Side("thin"), left=Side("thin"), right=Side("thin"))

# Same process for the third header
ws.append(["LONG TERM BLABLABLA"])
ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=5)
for cell in ws["3"]:
    cell.border = Border(bottom=Side("thin"), left=Side("thin"), right=Side("thin"))

# Here we resize every column except The first one ================================================================

ws.column_dimensions["B"].width = 10
ws.column_dimensions["C"].width = 10
ws.column_dimensions["D"].width = 10
ws.column_dimensions["E"].width = 10

# Now we iterate to create our rows from data ================================================================

for row in rows:
    # We resize the clumn for the first one for it to appear bigger
    ws.column_dimensions["A"].width = 50
    ws.append(
        total_fatigue_stress_styled_cells(
            ["Total fatigue stress in 1 year [-]", "", "", "", ""]
        )
    )
    # We resize the clumn for the first one for it to appear a bit smaller than the first generated row's first column
    ws.column_dimensions["A"].width = 30
    # Now we insert data with different styles
    ws.append(styled_cells(["Section", "Top 1", "Top 1", "Buoyancy", "Buoyancy"], True))
    ws.append(styled_cells(["Arc length [m]", row, row, row * 1000, row * 1000], True))
    ws.append(styled_cells(["Radial position", row, row, row * 1000, row * 1000], True))
    ws.append(styled_cells(["Angle", row, row, row * 1000, row * 1000], False))
    ws.append(styled_cells(["BLABLABLA", row, row, row * 1000, row * 1000], False))
    ws.append(styled_cells(["BLABLABLA", row, row, row * 1000, row * 1000], False))
    ws.append(bins_styled_cells(["BINS", row, row, row * 1000, row * 1000]))

# Save the file
wb.save("example.xlsx")
